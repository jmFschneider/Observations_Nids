"""
Commande Django pour charger les espèces d'oiseaux depuis la Liste des Oiseaux de France (LOF).

Cette commande télécharge et importe la Liste officielle des Oiseaux de France
publiée par la Commission de l'avifaune française (CAF) et disponible sur Faune-France.

Usage:
    python manage.py charger_lof
    python manage.py charger_lof --force
    python manage.py charger_lof --file /chemin/vers/LOF2024.xlsx
    python manage.py charger_lof --categories A,AC

Optimisé pour tous les environnements (Windows, Linux, Raspberry Pi).
"""

import gzip
import re
import shutil
from pathlib import Path

import openpyxl
import requests
from django.core.management.base import BaseCommand
from django.db import transaction

from taxonomy.models import Espece, Famille, Ordre


class Command(BaseCommand):
    help = "Charge les espèces d'oiseaux depuis la Liste des Oiseaux de France (LOF)"

    # URL de téléchargement de la LOF (Faune-France)
    LOF_URL = 'https://cdnfiles1.biolovision.net/www.faune-france.org/userfiles/FauneFrance/FFEnSavoirPlus/LOF2024IOC15.1032025.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--force',
            action='store_true',
            help='Force le rechargement même si des espèces existent déjà',
        )
        parser.add_argument(
            '--file',
            type=str,
            help='Chemin vers le fichier LOF Excel à importer (évite le téléchargement)',
        )
        parser.add_argument(
            '--categories',
            type=str,
            default='A,AC',
            help='Catégories à importer (séparées par des virgules). Défaut: A,AC (espèces sauvages)',
        )
        parser.add_argument(
            '--limit',
            type=int,
            help="Limite le nombre d'espèces à importer (pour tests)",
        )

    def handle(self, *args, **options):
        """Point d'entrée principal de la commande."""
        self.stdout.write(
            self.style.MIGRATE_HEADING('\n=== Chargement LOF - Oiseaux de France ===\n')
        )

        # Vérifier si déjà chargé
        especes_count = Espece.objects.filter(valide_par_admin=True).count()
        if especes_count > 100 and not options['force']:
            self.stdout.write(
                self.style.WARNING(
                    f'{especes_count} espèces déjà en base. '
                    'Utilisez --force pour recharger.'
                )
            )
            return

        # Obtenir le fichier LOF
        if options['file']:
            lof_file = Path(options['file'])
            if not lof_file.exists():
                self.stdout.write(self.style.ERROR(f"Fichier introuvable: {lof_file}"))
                return
            self.stdout.write(f"Utilisation du fichier local: {lof_file}")
        else:
            lof_file = self._download_lof()
            if not lof_file:
                return

        # Parser les catégories
        categories = [cat.strip().upper() for cat in options['categories'].split(',')]
        self.stdout.write(f"Catégories à importer: {', '.join(categories)}")

        # Supprimer les anciennes données si --force
        if options['force']:
            self._clear_database()

        # Importer les données
        stats = self._import_lof(lof_file, categories=categories, limit=options.get('limit'))

        # Afficher le rapport final
        self._display_report(stats)

    def _download_lof(self) -> Path | None:
        """Télécharge le fichier LOF depuis Faune-France."""
        self.stdout.write("Téléchargement de la Liste des Oiseaux de France...")
        self.stdout.write(f"URL: {self.LOF_URL}")

        # Créer le répertoire de téléchargement
        download_dir = Path('tmp/lof')
        download_dir.mkdir(parents=True, exist_ok=True)

        lof_file = download_dir / 'LOF2024.xlsx'
        lof_file_decompressed = download_dir / 'LOF2024_decompressed.xlsx'

        # Télécharger si nécessaire
        if not lof_file_decompressed.exists():
            try:
                response = requests.get(self.LOF_URL, stream=True, timeout=30)
                response.raise_for_status()

                # Téléchargement
                with open(lof_file, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)

                self.stdout.write(self.style.SUCCESS("[OK] Téléchargement terminé"))

                # Vérifier si le fichier est compressé ou déjà un Excel
                with open(lof_file, 'rb') as f:
                    magic = f.read(2)

                # PK = ZIP/XLSX, 1f8b = GZIP
                if magic == b'PK':
                    # Déjà un fichier Excel, pas de décompression nécessaire
                    self.stdout.write("Fichier Excel détecté (non compressé)")
                    shutil.move(str(lof_file), str(lof_file_decompressed))
                elif magic == b'\x1f\x8b':
                    # Fichier gzippé, décompresser
                    self.stdout.write("Décompression du fichier...")
                    with gzip.open(lof_file, 'rb') as f_in, open(lof_file_decompressed, 'wb') as f_out:
                        shutil.copyfileobj(f_in, f_out)
                    self.stdout.write(self.style.SUCCESS("[OK] Décompression terminée"))
                    lof_file.unlink()
                else:
                    self.stdout.write(self.style.ERROR(f"Format de fichier non reconnu: {magic!r}"))
                    return None

            except requests.RequestException as e:
                self.stdout.write(self.style.ERROR(f'Erreur lors du téléchargement: {e}'))
                return None
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Erreur lors de la décompression: {e}'))
                return None

        else:
            # Vérifier que le fichier en cache est valide
            try:
                with open(lof_file_decompressed, 'rb') as f:
                    magic = f.read(2)
                    if magic != b'PK':
                        # Fichier corrompu, le supprimer et re-télécharger
                        self.stdout.write(
                            self.style.WARNING(
                                "Fichier en cache corrompu, re-téléchargement..."
                            )
                        )
                        lof_file_decompressed.unlink()
                        return self._download_lof()  # Récursion pour re-télécharger

                self.stdout.write(f"Utilisation du fichier existant: {lof_file_decompressed}")
            except Exception:
                # Si erreur de lecture, supprimer et re-télécharger
                lof_file_decompressed.unlink(missing_ok=True)
                return self._download_lof()

        return lof_file_decompressed

    def _clear_database(self):
        """Supprime les données taxonomiques existantes."""
        self.stdout.write("Suppression des données existantes...")

        with transaction.atomic():
            especes_count = Espece.objects.count()
            familles_count = Famille.objects.count()
            ordres_count = Ordre.objects.count()

            Espece.objects.all().delete()
            Famille.objects.all().delete()
            Ordre.objects.all().delete()

            self.stdout.write(
                self.style.WARNING(
                    f"Supprimé: {especes_count} espèces, "
                    f"{familles_count} familles, {ordres_count} ordres"
                )
            )

    def _import_lof(self, lof_file: Path, categories: list[str], limit: int | None = None) -> dict:
        """
        Importe les données LOF depuis le fichier Excel.
        """
        self.stdout.write(f"\nImport des données depuis: {lof_file}")
        self.stdout.write(f"Catégories filtrées: {', '.join(categories)}\n")

        stats = {
            'total_lines': 0,
            'ordres_created': 0,
            'familles_created': 0,
            'especes_created': 0,
            'especes_skipped': 0,
            'erreurs': 0,
        }

        # Caches pour éviter les requêtes répétées
        ordres_cache: dict[str, Ordre] = {}
        familles_cache: dict[str, Famille] = {}

        # Variables de contexte
        current_ordre = None
        current_famille = None

        try:
            # Ouvrir le fichier Excel
            wb = openpyxl.load_workbook(lof_file)
            ws = wb['LOF IOC 11.1']  # Feuille principale

            self.stdout.write(f"Fichier ouvert: {ws.max_row} lignes à traiter")

            # Traiter ligne par ligne
            for row in ws.iter_rows(min_row=2, values_only=True):
                stats['total_lines'] += 1

                # S'assurer que les valeurs sont des chaînes de caractères
                col1, categorie, nom_sci, nom_fr = (
                    str(cell) if cell is not None else '' for cell in row[:4]
                )

                if not nom_sci:
                    continue

                nom_sci = nom_sci.strip()

                # Détection d'un ordre (tout en majuscules, pas de catégorie)
                if nom_sci.isupper() and not categorie:
                    ordre_nom = nom_sci.replace('\xa0', '').strip()
                    if ordre_nom and ordre_nom not in ordres_cache:
                        ordre, created = Ordre.objects.get_or_create(
                            nom=ordre_nom, defaults={'description': ''}
                        )
                        ordres_cache[ordre_nom] = ordre
                        current_ordre = ordre
                        if created:
                            stats['ordres_created'] += 1
                            self.stdout.write(f"  Ordre créé: {ordre_nom}")
                    continue

                # Détection d'une famille (finit par "dae")
                if nom_sci.endswith('dae\xa0') or nom_sci.endswith('dae'):
                    famille_nom = nom_sci.replace('\xa0', '').strip()
                    if famille_nom and current_ordre and famille_nom not in familles_cache:
                        famille, created = Famille.objects.get_or_create(
                            nom=famille_nom,
                            defaults={'ordre': current_ordre, 'description': ''},
                        )
                        familles_cache[famille_nom] = famille
                        current_famille = famille
                        if created:
                            stats['familles_created'] += 1
                            self.stdout.write(f"    Famille créée: {famille_nom}")
                    continue

                # Ignorer les sous-espèces (commencent par • ou espace)
                if nom_sci.startswith('•') or nom_sci.startswith(' '):
                    continue

                # Espèce valide (a une catégorie et un nom français)
                if categorie and nom_fr:
                    categorie = categorie.strip()

                    # Filtrer par catégorie
                    if categorie not in categories:
                        stats['especes_skipped'] += 1
                        continue

                    # Limiter si option --limit
                    if limit and stats['especes_created'] >= limit:
                        break

                    try:
                        # Nettoyer le nom scientifique (enlever l'autorité)
                        nom_sci_clean = self._clean_scientific_name(nom_sci)

                        # Créer l'espèce
                        espece, created = Espece.objects.get_or_create(
                            nom_scientifique=nom_sci_clean,
                            defaults={
                                'nom': nom_fr.strip(),
                                'nom_anglais': '',  # LOF n'a pas les noms anglais
                                'famille': current_famille,
                                'statut': categorie,
                                'valide_par_admin': True,
                                'commentaire': f'Import LOF 2024 - Catégorie {categorie}',
                            },
                        )

                        if created:
                            stats['especes_created'] += 1
                            if stats['especes_created'] % 50 == 0:
                                self.stdout.write(
                                    f"\rEspèces importées: {stats['especes_created']}", ending=''
                                )

                    except Exception as e:
                        stats['erreurs'] += 1
                        self.stdout.write(
                            self.style.WARNING(
                                f"\nErreur ligne {stats['total_lines']}: {e}\n"
                                f"  Espèce: {nom_fr} - {nom_sci}"
                            )
                        )

            self.stdout.write()  # Nouvelle ligne après la progression

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\nErreur lors de la lecture du fichier: {e}'))
            return stats

        return stats

    def _clean_scientific_name(self, nom_sci: str) -> str:
        """
        Nettoie le nom scientifique en enlevant l'autorité taxonomique.
        Ex: "Branta bernicla (Linnaeus, 1758)" -> "Branta bernicla"
        """
        # Enlever tout ce qui est entre parenthèses et après
        nom_sci = re.sub(r'\s*\([^)]*\).*$', '', nom_sci)
        # Enlever les espaces multiples
        nom_sci = re.sub(r'\s+', ' ', nom_sci)
        return nom_sci.strip()

    def _display_report(self, stats: dict):
        """Affiche le rapport final d'import."""
        self.stdout.write(self.style.MIGRATE_HEADING("\n=== Rapport d'import ===\n"))

        self.stdout.write(
            self.style.SUCCESS(
                f"Lignes traitées: {stats['total_lines']:,}\n"
                f"\nCréations:\n"
                f"   - Ordres: {stats['ordres_created']}\n"
                f"   - Familles: {stats['familles_created']}\n"
                f"   - Espèces: {stats['especes_created']}\n"
                f"   - Espèces ignorées (autres catégories): {stats['especes_skipped']}\n"
            )
        )

        if stats['erreurs'] > 0:
            self.stdout.write(self.style.WARNING(f"Erreurs: {stats['erreurs']}"))

        # Statistiques finales en base
        self.stdout.write(self.style.MIGRATE_HEADING('\n=== Base de données ===\n'))
        total_ordres = Ordre.objects.count()
        total_familles = Famille.objects.count()
        total_especes = Espece.objects.count()

        self.stdout.write(
            self.style.SUCCESS(
                f"Ordres: {total_ordres}\n"
                f"Familles: {total_familles}\n"
                f"Espèces: {total_especes}\n"
            )
        )

        # Exemples d'espèces
        self.stdout.write(self.style.MIGRATE_HEADING("Exemples d'espèces importées:"))
        exemples = Espece.objects.select_related('famille', 'famille__ordre')[:5]
        for esp in exemples:
            famille_info = f" ({esp.famille.nom})" if esp.famille else ""
            ordre_info = (
                f" - {esp.famille.ordre.nom}" if esp.famille and esp.famille.ordre else ""
            )
            self.stdout.write(f"  - {esp.nom}{famille_info}{ordre_info}")
            self.stdout.write(f"    {esp.nom_scientifique}")

        self.stdout.write(self.style.SUCCESS('\n[OK] Import termine avec succes!\n'))
